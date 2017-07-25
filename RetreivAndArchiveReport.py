"""
This program wil obtain archive and retrieve volume in MB for each ASP customer.
It will compile the data into worksheet and email it  to recipients.
"""
import pyodbc
import datetime
import smtplib
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders


def archive_retrieve_report():

    """
    This function will do 4 tasks:
    1.  Connect to SQL Server.
    2.  Run query and retrieve data from SQL server Data Base.
    3.  Write the data to Excel worksheet.
    4.  Email the worksheet to end user.
    """

    # File name for Excel Worksheet
    excel_filename = ""

    # Get all Virtual Archive names from SQL Server
    virtual_archives = get_archives()

    # Set up Excel Worksheet.
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "EA Monthly Report"
    work_sheet.append(['Archive Name', 'State', 'Data in MB', 'Data in GB', 'Data in TB'])

    # Assign font and background color properties for Column Title cells
    f = Font(name="Arial", size=14, bold=True, color="FF000000")
    fill = PatternFill(fill_type="solid", start_color="00FFFF00")

    # Format Worksheet columns
    for L in "ABCDE":
        work_sheet[L + "1"].fill = fill
        work_sheet[L + "1"].font = f
        if L in "AB":
            work_sheet.column_dimensions[L].width = 25.0
        if L in "CDE":
            work_sheet.column_dimensions[L].width = 35.0

    # Obtain Exam Volume for all virtual archives and write the data to excel sheet.
    for archive in virtual_archives:
        # Obtains archive volume from SQL server.
        rows = retrieve_volume(archive)

        for row in rows:

            # Adds archive name and exam volume to Worksheet.
            work_sheet.append([archive,                             # Archive Name
                               row[1],                              # State
                               round(row[2], 2),                    # Data in MB
                               round((row[2] / 1024), 2),           # Data in GB
                               round((row[2] / 1024 / 1024), 2)     # Data in TB
                               ])

        print("Added {} Archive and Retrieve Volume to Workbook!".format(archive))
        # Add blank line to worksheet for spacing
        work_sheet.append([])

    # Saves Excel worksheet.
    excel_filename = "ASP_Archive_Retrieve_Report_{}.xlsx".format(datetime.datetime.now().strftime("%Y-%m-%d"))
    work_book.save(excel_filename)

    # # Send email with attachment.
    # send_email(excel_filename)


# Obtain Archive Names from SQL Server.
def get_archives():
    """
    This function will return list of Virtual Archive Names from SQL Server.
    """
    # Create list to store archive names.
    archive_list = []

    # Obtain credentials from file
    with open("data.txt", "r") as f:
        read_data = f.readline().split()
        f.close()

    # Define data base connection parameters.
    sqlserver = 'SQL1'
    database = 'RSAdmin'
    username = read_data[0]
    password = read_data[1]

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER=' + sqlserver + ';'
        r'DATABASE=' + database + ';'
        r'UID=' + username + ';'
        r'PWD=' + password + ''
    )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                SELECT DBName from tblArchive
                ORDER BY DBName
                """)
    rows = cur.fetchall()

    # Add Archive names to archive list.
    for row in rows:
        archive_list.append(row[0])
    # Close SQL Connection.
    cur.close()
    conn.close()

    return archive_list


# Obtain archive volume.
def retrieve_volume(db_name):
    """
    This function will obtain archive volume form SQL server.
    """

    # Obtain credentials from file
    with open("data.txt", "r") as f:
        read_data = f.readline().split()
        f.close()

    # Define data base connection parameters.
    sqlserver = 'SQL1'
    database = db_name
    username = read_data[0]
    password = read_data[1]

    # Establish DB connections.
    conn = pyodbc.connect(
        r'DRIVER={SQL Server};'
        r'SERVER='+sqlserver+';'
        r'DATABASE='+database+';'
        r'UID='+username+';'
        r'PWD='+password+''
        )
    cur = conn.cursor()
    # Execute query on Data Base.
    cur.execute("""
                SELECT 
                  State,
                  CASE State 
                    When 1 Then 'Writable'
                    When 2 Then 'Read Only'
                    When 3 Then 'Frozen'
                    When 4 Then 'Archived'
                    When 5 Then 'Out Cache'
                    Else 'Unknown'
                  End AS "States",
                 -- Count(State) AS "# of Images",
                  Sum(Abs(Bytesize/1024/1024)) AS "MB of data",
                  GetDate() AS "Date/Time"
                from tblfile with (nolock)
                group by state with Rollup
                order by state      
                """)
    rows = cur.fetchall()
    # Close SQL Server Connections.
    cur.close()
    conn.close()
    return rows


# Send email with Report
def send_email(file_attachment):
    """This function will send email with the attachment.
    It takes attachment file name as argument.
    """

    # Define email body
    body = "This is Archive and Retrieve volume for all ASP Customers."
    content = MIMEText(body, 'plain')

    # Open file attachment
    filename = file_attachment
    infile = open(filename, "rb")

    # Set up attachment to be send in email
    part = MIMEBase("application", "octet-stream")
    part.set_payload(infile.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)

    msg = MIMEMultipart("alternative")

    # Define email recipients
    to_email = [
        "na@na.com"
        ]
    # Define From email
    from_email = "na@na.com"

    # Create email content
    msg["Subject"] = "ASP Archive Retrieve Report {}".format(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    msg["From"] = from_email
    msg["To"] = ",".format(to_email)
    msg.attach(part)
    msg.attach(content)
    # Send email to SMTP server
    s = smtplib.SMTP("10.4.1.1", 25)
    s.sendmail(from_email, to_email, msg.as_string())
    s.close()

# Run script
archive_retrieve_report()
